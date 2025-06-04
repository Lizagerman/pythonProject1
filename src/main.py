import json
import csv
import openpyxl
from datetime import datetime
from typing import List, Dict, Any, Optional

from src.transaction import search_transactions_by_description


def main():
    print("Привет, Liza! Проект жив 👋")


if __name__ == "__main__":
    main()


# Предполагаем, что эти функции
# (load_transactions_from_json, load_transactions_from_csv,
# load_transactions_from_xlsx, filter_by_status,
# sort_transactions_by_date,
# filter_by_currency, format_transaction) уже реализованы в других модулях
# или будут реализованы
# тобой. Для примера я могу дать их заглушки или простые реализации.


# Заглушки или простые реализации для демонстрации main
def load_transactions_from_json(filepath: str) -> List[Dict[str, Any]]:
    """Загружает транзакции из JSON файла."""
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Ошибка: Файл {filepath} не найден.")
        return []
    except json.JSONDecodeError:
        print(f"Ошибка: Некорректный JSON формат в файле {filepath}.")
        return []


def load_transactions_from_csv(filepath: str) -> List[Dict[str, Any]]:
    """Загружает транзакции из CSV файла."""
    transactions: List[Dict[str, Any]] = []
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                transactions.append(dict(row))
    except FileNotFoundError:
        print(f"Ошибка: Файл {filepath} не найден.")
    except Exception as e:
        print(f"Ошибка при чтении CSV файла: {e}")
    return transactions


def load_transactions_from_xlsx(filepath: str) -> List[Dict[str, Any]]:
    """Загружает транзакции из XLSX файла."""
    transactions: List[Dict[str, Any]] = []
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        for row_idx in range(2, sheet.max_row + 1):
            transaction = {}
            for col_idx, header in enumerate(headers):
                transaction[header] = sheet.cell(row=row_idx, column=col_idx + 1).value
            transactions.append(transaction)
    except FileNotFoundError:
        print(f"Ошибка: Файл {filepath} не найден.")
    except Exception as e:
        print(f"Ошибка при чтении XLSX файла: {e}")
    return transactions


def filter_by_status(
    transactions: List[Dict[str, Any]], status: str
) -> List[Dict[str, Any]]:
    """Фильтрует транзакции по статусу."""
    return [t for t in transactions if t.get("state", "").lower() == status.lower()]


def sort_transactions_by_date(
    transactions: List[Dict[str, Any]], ascending: bool = True
) -> List[Dict[str, Any]]:
    """Сортирует транзакции по дате."""
    try:
        return sorted(
            transactions,
            key=lambda x: datetime.strptime(x.get("date", ""), "%Y-%m-%dT%H:%M:%S.%f"),
            reverse=not ascending,
        )
    except (ValueError, TypeError):
        return transactions  # Возвращаем несортированный список в случае ошибки даты


def filter_by_currency(
    transactions: List[Dict[str, Any]], currency_code: str = "RUB"
) -> List[Dict[str, Any]]:
    """Фильтрует транзакции по валюте."""
    return [
        t
        for t in transactions
        if t.get("operationAmount", {}).get("currency", {}).get("code", "").upper()
        == currency_code.upper()
    ]


def format_transaction(transaction: Dict[str, Any]) -> str:
    """Форматирует информацию о транзакции для вывода."""
    date_str = ""
    try:
        date_obj = datetime.strptime(
            transaction.get("date", ""), "%Y-%m-%dT%H:%M:%S.%f"
        )
        date_str = date_obj.strftime("%d.%m.%Y")
    except (ValueError, TypeError):
        pass  # Игнорируем ошибки форматирования даты

    description = transaction.get("description", "Нет описания")
    amount = transaction.get("operationAmount", {}).get("amount", "N/A")
    currency = (
        transaction.get("operationAmount", {}).get("currency", {}).get("name", "")
    )

    from_account = transaction.get("from", "")
    to_account = transaction.get("to", "")

    # Обрезаем номера счетов для анонимности
    if from_account:
        if len(from_account) > 4:
            from_account = (
                f"{from_account[:-10]} **{from_account[-4:]}"  # Пример маскировки
            )
        else:
            from_account = f"Счет **{from_account[-4:]}"  # Пример маскировки
    if to_account:
        if len(to_account) > 4:
            to_account = f"{to_account[:-10]} **{to_account[-4:]}"  # Пример маскировки
        else:
            to_account = f"Счет **{to_account[-4:]}"  # Пример маскировки

    # Если есть "from" и "to"
    if from_account and to_account:
        accounts_info = f"{from_account} -> {to_account}"
    elif to_account:  # Если только "to"
        accounts_info = to_account
    else:
        accounts_info = ""  # Если нет информации об аккаунтах

    return (
        f"{date_str} {description}\n"
        f"{accounts_info}\n"
        f"Сумма: {amount} {currency}\n"
    )


def main() -> None:
    """
    Основная логика программы, предоставляющая пользовательский интерфейс для работы
    с банковскими транзакциями.
    """
    print("Привет! Добро пожаловать в программу работы с банковскими транзакциями.")
    transactions: List[Dict[str, Any]] = []

    while True:
        print("\nВыберите необходимый пункт меню:")
        print("1. Получить информацию о транзакциях из JSON-файла")
        print("2. Получить информацию о транзакциях из CSV-файла")
        print("3. Получить информацию о транзакциях из XLSX-файла")
        print("0. Выход")

        choice = input("Пользователь: ")

        if choice == "1":
            print("Программа: Для обработки выбран JSON-файл.")
            filepath = input("Введите путь к JSON файлу (например, operations.json): ")
            transactions = load_transactions_from_json(filepath)
            if not transactions:
                print("Не удалось загрузить транзакции. Попробуйте снова.")
                continue
            break
        elif choice == "2":
            print("Программа: Для обработки выбран CSV-файл.")
            filepath = input("Введите путь к CSV файлу (например, operations.csv): ")
            transactions = load_transactions_from_csv(filepath)
            if not transactions:
                print("Не удалось загрузить транзакции. Попробуйте снова.")
                continue
            break
        elif choice == "3":
            print("Программа: Для обработки выбран XLSX-файл.")
            filepath = input("Введите путь к XLSX файлу (например, operations.xlsx): ")
            transactions = load_transactions_from_xlsx(filepath)
            if not transactions:
                print("Не удалось загрузить транзакции. Попробуйте снова.")
                continue
            break
        elif choice == "0":
            print("Программа: До свидания!")
            return
        else:
            print("Программа: Некорректный выбор. Пожалуйста, введите 1, 2, 3 или 0.")

    filtered_transactions = list(
        transactions
    )  # Создаем копию для дальнейшей фильтрации

    # Фильтрация по статусу
    available_statuses = ["EXECUTED", "CANCELED", "PENDING"]
    while True:
        status_input = input(
            "Программа: Введите статус, по которому необходимо выполнить фильтрацию. "
            f"Доступные для фильтровки статусы: {', '.join(available_statuses)}\n"
            "Пользователь: "
        )
        if status_input.upper() in available_statuses:
            filtered_transactions = filter_by_status(
                filtered_transactions, status_input
            )
            print(
                f'Программа: Операции отфильтрованы по статусу "{status_input.upper()}"'
            )
            break
        else:
            print(f'Программа: Статус операции "{status_input}" недоступен.')

    # Сортировка по дате
    sort_by_date_input = input(
        "Программа: Отсортировать операции по дате? Да/Нет\nПользователь: "
    )
    if sort_by_date_input.lower() == "да":
        while True:
            sort_order_input = input(
                "Программа: Отсортировать по возрастанию "
                "или по убыванию?\nПользователь: "
            )
            if sort_order_input.lower() == "по возрастанию":
                filtered_transactions = sort_transactions_by_date(
                    filtered_transactions, ascending=True
                )
                break
            elif sort_order_input.lower() == "по убыванию":
                filtered_transactions = sort_transactions_by_date(
                    filtered_transactions, ascending=False
                )
                break
            else:
                print(
                    "Программа: Некорректный ввод. Пожалуйста, "
                    "введите 'по возрастанию' или 'по убыванию'."
                )

    # Фильтрация по рублевым транзакциям
    rub_only_input = input(
        "Программа: Выводить только рублевые транзакции? Да/Нет\nПользователь: "
    )
    if rub_only_input.lower() == "да":
        filtered_transactions = filter_by_currency(filtered_transactions, "RUB")

    # Фильтрация по слову в описании (используем ранее созданную функцию)
    filter_by_description_input = input(
        "Программа: Отфильтровать список транзакций "
        "по определенному слову в описании? Да/Нет\nПользователь: "
    )
    if filter_by_description_input.lower() == "да":
        search_term = input(
            "Программа: Введите слово для поиска в описании:\nПользователь: "
        )
        filtered_transactions = search_transactions_by_description(
            filtered_transactions, search_term
        )

    print("\nПрограмма: Распечатываю итоговый список транзакций...")

    if not filtered_transactions:
        print(
            "Программа: Не найдено ни одной транзакции, "
            "подходящей под ваши условия фильтрации"
        )
    else:
        print(f"\nВсего банковских операций в выборке: {len(filtered_transactions)}\n")
        for transaction in filtered_transactions:
            print(format_transaction(transaction))


if __name__ == "__main__":
    main()

